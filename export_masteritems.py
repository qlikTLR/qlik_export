#pip install aiohttp asyncio dotenv websockets openpyxl
# written by Thomas Lindackers

import openpyxl # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from classes.qlik_wrapper import QlikWrapper
from dotenv import load_dotenv # type: ignore
import asyncio, os  
from classes.terminal_helper import TerminalHelper
from classes.file_writer_helper import FileWriterHelper
load_dotenv()
API_KEY = os.getenv("API_KEY")
APP_ID = "37e72b96-b3f2-445b-89a8-d79c90c965f3"
QLIK_TENANT = "qlikinternal.us.qlikcloud.com"

def sanitize_excel_value(value, protect_formula=True):
    """
    Converts value to string and protects against Excel interpreting it as a formula.
    """
    val = str(value)
    if protect_formula and val.startswith(("=", "+", "-", "/", "*")):
        return "'" + val
    return val   


async def export_masteritems():
    qlik = QlikWrapper(QLIK_TENANT, API_KEY)
    APP_ID = input('Please enter a App ID for the export: ')
    await qlik.connect(APP_ID)

    app_info = await qlik.get_app_info(APP_ID)
   
    TerminalHelper.print_single_array(app_info, 'Start Export for','  ') 
    app_master_info = await qlik.get_masteritems_detailed()

    app_name = app_info.get("name", "NoAppName")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "App Info"

    for idx, (key, value) in enumerate(app_info.items(), start=1):
        ws1[f"A{idx}"] = key
        ws1[f"B{idx}"] = value

    ws2 = wb.create_sheet(title="Masteritems")
    
    columns = ["ID", "Title", "Description", "Expression", "Tags", "Type", "ItemType"]
    for col_num, column_title in enumerate(columns, start=1):
        col_letter = get_column_letter(col_num)
        ws2[f"{col_letter}1"] = column_title  

    for row_num, item in enumerate(app_master_info, start=2):
        # Replace special characters in all fields except ID (item[0]) and Title (item[1])
       
        # Write values to sheet
        ws2[f"A{row_num}"] = item[0]  # ID
        ws2[f"B{row_num}"] = item[1]  # Title
        ws2[f"C{row_num}"] = sanitize_excel_value(item[2])  # Description
        ws2[f"D{row_num}"] = sanitize_excel_value(item[3])  # Expression
        ws2[f"E{row_num}"] = sanitize_excel_value(', '.join(str(tag) for tag in item[4]) if isinstance(item[4], list) else item[4])
        ws2[f"F{row_num}"] = sanitize_excel_value(item[5])  # Type
        ws2[f"G{row_num}"] = sanitize_excel_value(item[6])  # ItemType

    
    excel_filename = f"exports/{app_name} MasterItems.xlsx"
    wb.save(excel_filename)

    print(f"   exported to:  {excel_filename} ")

    await qlik.close()

if __name__ == "__main__":
    asyncio.run(export_masteritems())
